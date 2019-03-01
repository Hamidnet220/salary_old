from models import *
from openpyxl import load_workbook


wb=load_workbook('/mnt/work/arvandan/employeeData.xlsx')
ws=wb.get_sheet_by_name('employees')



#for i in range(2,438):
fname =ws.cell(row = 2 , column = 3).value
lname =ws.cell(row = 2 , column = 4).value
fathername =ws.cell(row = 2 , column = 5).value
ncode = ws.cell(row = 2 , column = 2).value 
id_num = ws.cell(row = 2 , column = 6).value 
bank_acc =ws.cell(row =2 , column =18).value
inscuranc =ws.cell(row =2 , column =33).value

emp=Employee(firstname=fname , lastname=lname , father_name = fathername,national_id= ncode,
        is_maride=True,status=EmployeeStatus.pk.filter(id__id=1),id_number=id_num,
        bank_account = bank_acc,insurance_id=inscuranc)  

